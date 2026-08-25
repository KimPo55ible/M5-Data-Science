"""
ETL pipeline: Care Leaver Housing Outcomes (Manchester vs England)
Data Science Project, M5 Data Science Professional Practice

IMPORTANT - REPRODUCIBILITY NOTE:
This script was built and tested against source files downloaded on 22/08/2026.
Public data catalogues (DfE Explore Education Statistics, MHCLG) are periodically
revised, so re-downloading these files today may not reproduce the exact figures
reported in the accompanying write-up. To reproduce the original analysis exactly,
use the source files as published/accessed on 22/08/2026 (see References in the
report for direct links).

Required input files (place in the same folder as this script):
    1. la_care_leavers_accommodation_suitability 2021 - 2025.csv   (DfE)
    2. Statutory_Homelessness_England_Time_Series_Live.xlsx        (MHCLG)
    3. Statutory_Homelessness_Detailed_Local_Authority_Data_2023-2024_Revised.xlsx   (MHCLG)
    4. Copy of Statutory_Homelessness_Detailed_Local_Authority_Data_2024-2025_corrected.xlsx (MHCLG)

Run with:
    pip install pandas openpyxl
    python Analysis.py

Output:
    master_joined_table.csv
"""

import pandas as pd
import openpyxl

# --- Source 1: DfE accommodation suitability, 19-21 age band ---
df1 = pd.read_csv('la_care_leavers_accommodation_suitability 2021 - 2025.csv')

acc = df1[(df1['care_leaver_age'] == '19 to 21 years') &
          (df1['breakdown'] == 'Accommodation considered unsuitable') &
          ((df1['la_name'] == 'Manchester') |
           ((df1['la_name'].isna()) & (df1['geographic_level'] == 'National')))].copy()

acc['area'] = acc['la_name'].fillna('England')
acc = acc[['time_period', 'area', 'care_leaver_count', 'care_leaver_percent']].rename(
    columns={'time_period': 'year',
             'care_leaver_count': 'accom_unsuitable_count',
             'care_leaver_percent': 'accom_unsuitable_pct'})


# --- Source 2a: MHCLG homelessness, England national time series, 2018-19 to 2024-25 ---
def get_england_time_series(path, sheet):
    """Pull the annual England row for each year from the time series workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=7, max_row=13, values_only=True):
        year_label = row[0].strip() if row[0] else None
        if year_label and 'to March' in year_label:
            end_year = int(year_label.split(' ')[-1])
            rows.append({
                'year': end_year,
                'area': 'England',
                'homeless_count': row[14],   # Care leaver aged 18-20 support need
                'total_owed': row[3],        # Total households owed a duty
            })
    return pd.DataFrame(rows)


nat_home = get_england_time_series('Statutory_Homelessness_England_Time_Series_Live.xlsx', 'A3')
nat_home['homeless_pct_of_all'] = (nat_home['homeless_count'] / nat_home['total_owed'] * 100).round(2)


# --- Source 2b: MHCLG homelessness, Manchester, care leaver 18-20 support-need flag ---
def get_la_row(path, sheet, la_name):
    """Search a workbook sheet for the row matching a given local authority name."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if row[1] and la_name in str(row[1]):
            return row


manc_2425 = get_la_row(
    'Copy of Statutory_Homelessness_Detailed_Local_Authority_Data_2024-2025_corrected.xlsx',
    'A3', 'Manchester')
manc_2324 = get_la_row(
    'Statutory_Homelessness_Detailed_Local_Authority_Data_2023-2024_Revised.xlsx',
    'A3', 'Manchester')

# Column 16 = "Care leaver aged 18-20 years" support need count
# Columns 4, 5, 7 = households with no support needs, unknown, and one-or-more support needs
manc_home = pd.DataFrame([
    {'year': 2024, 'area': 'Manchester',
     'homeless_count': manc_2324[16],
     'total_owed': manc_2324[4] + manc_2324[5] + manc_2324[7]},
    {'year': 2025, 'area': 'Manchester',
     'homeless_count': manc_2425[16],
     'total_owed': manc_2425[4] + manc_2425[5] + manc_2425[7]},
])
manc_home['homeless_pct_of_all'] = (manc_home['homeless_count'] / manc_home['total_owed'] * 100).round(2)

home = pd.concat([nat_home, manc_home], ignore_index=True)

# --- Join on shared identifiers: year and area ---
master = acc.merge(home, on=['year', 'area'], how='outer').sort_values(['area', 'year'])
master.to_csv('master_joined_table.csv', index=False)

print(master.to_string(index=False))